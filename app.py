import streamlit as st
import pandas as pd
import datetime
import re
import io
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment

# --- ページ基本設定 ---
st.set_page_config(page_title="Keepa 統合管理ツール", layout="centered")

# --- サイドバー・メニュー ---
st.sidebar.title("🛠 ツール選択")
mode = st.sidebar.radio(
    "使用する機能を選んでください",
    ["keepaデータ見積作成ツール", "リサーチ表用keepaデータ加工ツール"]
)

# --- 共通関数：カッコ内のテキストを削除 ---
def clean_text(text):
    if pd.isna(text):
        return text
    # ( ) や （ ） とその中身を削除する正規表現
    cleaned = re.sub(r'[\(（].*?[\)）]', '', str(text))
    return cleaned.strip()

# ==========================================
# 1. keepaデータ見積作成ツール（旧：見積作成ツール）
# ==========================================
if mode == "keepaデータ見積作成ツール":
    st.title("📦 keepaデータ見積作成ツール")
    st.info("画像・商品名・ASIN・EANの抽出と、黒枠を適用します。")
    
    uploaded_file = st.file_uploader("エクセルファイルをアップロード", type=["xlsx"], key="quote_uploader")

    if uploaded_file:
        try:
            df = pd.read_excel(uploaded_file)
            # 列名の前後スペースを削除
            df.columns = df.columns.astype(str).str.strip()
            
            new_data = {}
            for col in df.columns:
                c_low = str(col).lower()
                if ('image' in c_low or '画像' in c_low) and '画像' not in new_data:
                    new_data['画像'] = df[col]
                elif ('title' in c_low or '商品名' in c_low) and '商品名' not in new_data:
                    new_data['商品名'] = df[col]
                elif 'asin' == c_low and 'ASIN' not in new_data:
                    new_data['ASIN'] = df[col]
                elif 'ean' in c_low and 'EAN' not in new_data:
                    new_data['EAN'] = df[col].apply(lambda x: '{:.0f}'.format(x) if pd.notnull(x) and isinstance(x, (int, float)) else str(x) if pd.notnull(x) else "")

            df_filtered = pd.DataFrame(new_data)
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_filtered.to_excel(writer, index=False)
            
            output.seek(0)
            wb = load_workbook(output)
            ws = wb.active
            
            side = Side(style='thin', color='000000')
            border = Border(top=side, bottom=side, left=side, right=side)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')

            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20

            final_output = io.BytesIO()
            wb.save(final_output)
            
            st.success("変換準備ができました！")
            st.download_button(
                label="📥 見積書をダウンロード",
                data=final_output.getvalue(),
                file_name=f"{datetime.datetime.now().strftime('%y%m%d')}_見積書.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"エラーが発生しました: {e}")

# ==========================================
# 2. リサーチ表用keepaデータ加工ツール（旧：データ加工ツール）
# ==========================================
elif mode == "リサーチ表用keepaデータ加工ツール":
    st.title("📊 リサーチ表用keepaデータ加工ツール")
    st.info("1枚目は履歴として保持し、2枚目でD列(製造者)のクリーンアップと重複削除を行います。")

    uploaded_file = st.file_uploader("Keepaのエクセルファイルをアップロード", type=["xlsx"], key="process_uploader")

    if uploaded_file:
        try:
            # 1. 元データの読み込み
            df_all = pd.read_excel(uploaded_file, sheet_name=0)
            # 列名の前後にある空白を削除
            df_all.columns = df_all.columns.astype(str).str.strip()
            
            date_match = re.search(r'\d{4}-\d{2}-\d{2}', uploaded_file.name)
            sheet1_name = date_match.group(0) if date_match else "履歴データ"

            # 2. 2枚目用のデータ加工
            target_cols = ['商品名', '売れ筋ランキング: 現在価格', 'ASIN', '製造者', 'ブランド']
            existing_cols = [c for c in target_cols if c in df_all.columns]
            df_processed = df_all[existing_cols].copy()
            
            # 3. D列（製造者）のクリーンアップ
            if '製造者' in df_processed.columns:
                df_processed['製造者'] = df_processed['製造者'].apply(clean_text)
            
            # 4. D列（製造者）を基準に重複削除
            if '製造者' in df_processed.columns:
                before_count = len(df_processed)
                df_processed = df_processed.drop_duplicates(subset=['製造者'], keep='first')
                after_count = len(df_processed)
            else:
                st.warning("列名に『製造者』が見つかりませんでした。")

            # 5. ダウンロード用エクセル作成
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_all.to_excel(writer, sheet_name=sheet1_name, index=False)
                df_processed.to_excel(writer, sheet_name="重複削除", index=False)
            
            st.success(f"加工完了！重複削除により {before_count}件 → {after_count}件 になりました。")
            st.download_button(
                label="📥 加工済みデータをダウンロード",
                data=output.getvalue(),
                file_name=f"Processed_{uploaded_file.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"エラーが発生しました: {e}")